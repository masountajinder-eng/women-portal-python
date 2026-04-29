from flask import Flask, render_template, request, jsonify, send_file, redirect, session
import os
from dotenv import load_dotenv
import random
import base64
from openpyxl import Workbook, load_workbook
from threading import Lock
from datetime import datetime, timedelta
import requests

import smtplib
from email.mime.multipart import MIMEMultipart
from email.mime.text import MIMEText
from email.mime.base import MIMEBase
from email import encoders

# ================= INIT =================
load_dotenv()

app = Flask(__name__, static_folder='static', template_folder='templates')
app.secret_key = os.getenv("SECRET_KEY", "secret123")

print("🔥 EMAIL SYSTEM RUNNING")

# ================= ADMIN =================
ADMIN_USER = "237engrregt"
ADMIN_PASS = "237237chakde"

# ================= GOOGLE SHEET =================
GOOGLE_SCRIPT_URL = "https://script.google.com/macros/s/AKfycbx0rXqDF0ZiMoSMUyF3VhHELxnlNydD-ACBFir42EqwaJkSZ5u7IqK8S3I63scwHlUj1w/exec"

# ================= RESEND EMAIL =================
def send_alert_email(data):
    try:
        api_key = os.getenv("RESEND_API_KEY")

        attachments = []
        if data.get("audio") and os.path.exists(data["audio"]):
            with open(data["audio"], "rb") as f:
                encoded_file = base64.b64encode(f.read()).decode()

            attachments.append({
                "filename": os.path.basename(data["audio"]),
                "content": encoded_file
            })

        response = requests.post(
            "https://api.resend.com/emails",
            headers={
                "Authorization": f"Bearer {api_key}",
                "Content-Type": "application/json"
            },
            json={
                "from": "Women Sahara Portal <onboarding@resend.dev>",
                "to": ["237engrregt@gmail.com"],
                "subject": f"Complaint Received - {data['complaint_id']}",
                "text": f"New complaint from {data['name']}",
                "html": f"<h3>{data['complaint']}</h3>",
                "attachments": attachments
            }
        )

        print("📧 RESEND STATUS:", response.status_code)

    except Exception as e:
        print("❌ RESEND ERROR:", e)

# ================= GMAIL =================
def send_fallback_gmail(data):
    try:
        user = os.getenv("SMTP_USER")
        password = os.getenv("SMTP_PASS")

        msg = MIMEMultipart()
        msg["From"] = user
        msg["To"] = user
        msg["Subject"] = f"Complaint {data['complaint_id']}"

        msg.attach(MIMEText(data["complaint"], "plain"))

        if data.get("audio") and os.path.exists(data["audio"]):
            part = MIMEBase('application', 'octet-stream')
            with open(data["audio"], "rb") as f:
                part.set_payload(f.read())
            encoders.encode_base64(part)
            part.add_header('Content-Disposition',
                            f'attachment; filename="{os.path.basename(data["audio"])}"')
            msg.attach(part)

        server = smtplib.SMTP("smtp.gmail.com", 587)
        server.starttls()
        server.login(user, password)
        server.send_message(msg)
        server.quit()

        print("📧 GMAIL SENT")

    except Exception as e:
        print("❌ GMAIL ERROR:", e)

# ================= GOOGLE SHEET =================
def send_to_google_sheet(data):
    try:
        payload = {
            "complaint_id": data["complaint_id"],
            "name": data["name"],
            "contact": data["contact"],
            "complaint": data["complaint"],
            "category": data["category"],
            "subcategory": data["subcategory"]
        }

        res = requests.post(GOOGLE_SCRIPT_URL, json=payload, timeout=10)

        print("📊 SHEET STATUS:", res.status_code)
        print("📊 SHEET RESPONSE:", res.text)

    except Exception as e:
        print("❌ SHEET ERROR:", e)

# ================= FILE =================
UPLOAD_FOLDER = "/tmp/audio_files"
os.makedirs(UPLOAD_FOLDER, exist_ok=True)

EXCEL_FILE = "/tmp/complaints.xlsx"
excel_lock = Lock()

def create_excel():
    if not os.path.exists(EXCEL_FILE):
        wb = Workbook()
        ws = wb.active
        ws.append([
            "ID","Complaint ID","Name","Address","Contact",
            "Email","Unit","WO","Quarter","Complaint",
            "Category","Subcategory","Reply","Audio","Date"
        ])
        wb.save(EXCEL_FILE)

create_excel()

def save_to_excel(data):
    with excel_lock:
        wb = load_workbook(EXCEL_FILE)
        ws = wb.active
        now = datetime.now().strftime("%Y-%m-%d %H:%M:%S")

        ws.append([
            ws.max_row,
            data["complaint_id"],
            data["name"],
            data["address"],
            data["contact"],
            data["email"],
            data["unit"],
            data["wo"],
            data["quarter"],
            data["complaint"],
            data["category"],
            data["subcategory"],
            data["reply"],
            data["audio"],
            now
        ])
        wb.save(EXCEL_FILE)

# ================= ROUTES =================

@app.route('/')
def home():
    return render_template("landing.html")

@app.route('/login', methods=['GET','POST'])
def login():
    if request.method == 'POST':
        if request.form.get("username") == ADMIN_USER and request.form.get("password") == ADMIN_PASS:
            session['admin'] = True
            return redirect("/admin")
        return render_template("login.html", error="❌ Wrong credentials")
    return render_template("login.html")

@app.route('/logout')
def logout():
    session.clear()
    return redirect("/login")

@app.route('/complaint', methods=['GET','POST'])
def complaint():
    if request.method == 'POST':
        try:
            complaint_id = "CMP" + str(random.randint(10000,99999))

            data = {
                "complaint_id": complaint_id,
                "name": request.form.get('name',''),
                "address": request.form.get('address',''),
                "contact": request.form.get('contact',''),
                "email": request.form.get('email',''),
                "unit": request.form.get('unit',''),
                "wo": request.form.get('wo',''),
                "quarter": request.form.get('quarter',''),
                "complaint": request.form.get('complaint',''),
                "category": request.form.get('category',''),
                "subcategory": request.form.get('subcategory',''),
                "reply": "Pending",
                "audio": ""
            }

            audio_data = request.form.get("audio_data")
            if audio_data and "," in audio_data:
                header, encoded = audio_data.split(",",1)
                filepath = os.path.join(UPLOAD_FOLDER, f"{complaint_id}.webm")
                with open(filepath,"wb") as f:
                    f.write(base64.b64decode(encoded))
                data["audio"] = filepath

            save_to_excel(data)

            # 🔥 FINAL FIX (CRASH PROOF)
            try:
                send_alert_email(data)
            except Exception as e:
                print("EMAIL ERROR:", e)

            try:
                send_fallback_gmail(data)
            except Exception as e:
                print("GMAIL ERROR:", e)

            try:
                send_to_google_sheet(data)
            except Exception as e:
                print("SHEET ERROR:", e)

            return jsonify({"status":"success","id":complaint_id})

        except Exception as e:
            print("❌ MAIN ERROR:", e)
            return jsonify({"status":"error","message":str(e)})

    return render_template("complaint.html")

@app.route('/admin')
def admin():
    if not session.get('admin'):
        return redirect("/login")

    data = []
    try:
        res = requests.get(GOOGLE_SCRIPT_URL, timeout=10)
        data = res.json()
    except:
        pass

    return render_template("admin.html", data=data)

@app.route('/track', methods=['GET','POST'])
def track():
    result = None

    if request.method == 'POST':
        cid = request.form.get("complaint_id")

        try:
            res = requests.get(GOOGLE_SCRIPT_URL, timeout=10)
            sheet_data = res.json()

            for row in sheet_data[1:]:
                if row[0] == cid:
                    result = row
                    break
        except:
            pass

    return render_template("track.html", result=result)

@app.route('/download_excel')
def download_excel():
    try:
        wb = load_workbook(EXCEL_FILE)
        ws = wb.active

        new_wb = Workbook()
        new_ws = new_wb.active

        new_ws.append([cell.value for cell in ws[1]])

        now = datetime.now()
        last_24 = now - timedelta(hours=24)

        for row in ws.iter_rows(min_row=2, values_only=True):
            try:
                row_time = datetime.strptime(row[14], "%Y-%m-%d %H:%M:%S")
                if row_time >= last_24:
                    new_ws.append(row)
            except:
                continue

        file_path = "/tmp/last24h.xlsx"
        new_wb.save(file_path)

        return send_file(file_path, as_attachment=True)

    except Exception as e:
        print("❌ DOWNLOAD ERROR:", e)
        return "Error"

if __name__ == "__main__":
    app.run(host="0.0.0.0", port=10000)